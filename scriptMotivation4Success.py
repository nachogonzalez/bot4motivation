import json
import tweepy
import random
from datetime import date
from datetime import datetime
from openpyxl import load_workbook
from loguru import logger

# Twitter authentication using tweepy
apiKey = "FeCDkJOLyYUNLxoRDR7hYHS0a"
apiSecretKey = "E4TUCiZqToRKLQqeIZN7Jwo7wMdQH36Ip08Up7agSYVMb32a0i"
accessToken = "1387362642826829825-CrTIqG1Dyq7xp6TEUSOgZ6LQ4NLUIt"
accessTokenSecret = "1pHHRlXxpjm2ne2kI5z7ohSIPkcRDbMGxuGn3ppUprzAD"

auth = tweepy.OAuthHandler(apiKey, apiSecretKey)
auth.set_access_token(accessToken, accessTokenSecret)

# Creation of the API object
api = tweepy.API(auth)

# Get data from xlsx file with all the quotes and feed the list
quotesList = []
wb = load_workbook(filename = './MotivationalQuotesv21.xlsx')
sheet_ranges = wb['Sheet1']
maxRows = sheet_ranges.max_row
indexLetter = "A"
indexNumber = 1
index = indexLetter + str(indexNumber)

while indexNumber < maxRows:
	quotesList.append(sheet_ranges[index].value)
	indexNumber = indexNumber + 1
	index = indexLetter + str(indexNumber)
	
# Get a random quote
randomIndex = random.randrange(maxRows)

# Build the tweet
tweet = (quotesList[randomIndex] + "\n #Motivation4Success #inspiration #luxury #lifestyle #enterpreneurship #hardwork #success #quotes #business")

# Post the tweet
api.update_status(tweet)

# Show logs on console
logger.info("scriptMotivation4Success: " + str(datetime.now()))
logger.info("Number of quotes available: " + str(maxRows))
logger.info("Random quote: " + str(randomIndex))
logger.info("Tweet posted")
logger.info("------------")
logger.info(tweet)